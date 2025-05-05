"""
BOM Comparison Tool
This script provides a PyQt5-based GUI application for comparing two Bill of Materials (BOM) Excel files.
It allows users to select two BOM files, apply various filters, and generate a comparison report in Excel format.
Classes:
    BOMComparisonApp(QMainWindow): Main application window for the BOM comparison tool.
Functions:
    __init__(): Initializes the main application window.
    init_ui(): Sets up the user interface components.
    create_button(text, callback): Creates a QPushButton with a specified callback function.
    select_file(label, file_attr): Generic method to select a file and update the corresponding label.
    select_file1(): Opens a file dialog to select the first BOM file.
    select_file2(): Opens a file dialog to select the second BOM file.
    run_comparison(): Executes the BOM comparison process and generates the output files.
    get_output_file_name(): Generates the output file name based on the selected BOM files.
    read_excel_files(): Reads the selected BOM Excel files into pandas DataFrames.
    merge_and_filter_dataframes(df1, df2): Merges and applies filters to the BOM DataFrames.
    handle_existing_file(output_file): Handles cases where the output file already exists.
    write_to_excel(output_file, df1, df2, df_merged): Writes the comparison results to an Excel file.
    handle_permission_error(): Handles permission errors when writing to a file.
    modify_excel_file(output_file): Modifies the Excel file using openpyxl to add formatting and filters.
    write_refdes_to_txt(df_merged): Writes the RefDes list from the merged DataFrame to a text file.
Constants:
    COLUMNS: List of column names to be used when reading the BOM Excel files.
    OUTPUT_DIR: Directory where the output files will be saved.
    EXCLUDED_PREFIX: Prefix used to exclude certain files.
Usage:
    Run the script to launch the GUI application. Use the interface to select two BOM files, apply filters, and generate a comparison report.
"""
import os
import sys
import pandas as pd
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QVBoxLayout, QLabel, QPushButton, QFileDialog, QWidget, QMessageBox, QCheckBox
)
from PyQt5.QtGui import QIcon
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import time

# Constants
COLUMNS = [
    'RefDes', 'Part Number', 'Description', 'Producer Abbreviation',
    'Producer Part Name', 'Validity Flag', 'Validity Flag Remarks',
    'Case Size', 'Collective Numbers', 'Quantity', 'Placement Side Information'
]
OUTPUT_DIR = 'Result'

class BOMComparisonApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("BOM Comparison Tool")
        self.setWindowIcon(QIcon("bosch.ico"))  # Set the application icon
        screen = QApplication.primaryScreen().availableGeometry()
        width, height = 600, 200  # Configurable size
        x = (screen.width() - width) // 2 # number 2 here means,
        y = (screen.height() - height) // 2
        self.setGeometry(x, y, width, height)

        self.file1 = None
        self.file2 = None

        self.init_ui()

    def init_ui(self):
        """Initialize the user interface."""
        layout = QVBoxLayout()

        self.label_file1 = QLabel("First BOM File: Not Selected")
        self.label_file2 = QLabel("Second BOM File: Not Selected")
        layout.addWidget(self.label_file1)
        layout.addWidget(self.label_file2)

        layout.addWidget(self.create_button("Select First BOM File", self.select_file1))
        layout.addWidget(self.create_button("Select Second BOM File", self.select_file2))

        self.checkbox_remove_zero_qty = QCheckBox("Remove Zero Quantity in Both BOMs")
        self.checkbox_remove_new_zero_qty = QCheckBox("Remove New Zero Quantity in BOM2")
        self.checkbox_remove_deprecated_zero_qty = QCheckBox("Remove Deprecated Zero Quantity in BOM1")
        self.checkbox_remove_unchanged = QCheckBox("Remove Unchanged Items")

        for checkbox in [
            self.checkbox_remove_zero_qty,
            self.checkbox_remove_new_zero_qty,
            self.checkbox_remove_deprecated_zero_qty,
            self.checkbox_remove_unchanged
        ]:
            layout.addWidget(checkbox)

        layout.addWidget(self.create_button("Run Comparison", self.run_comparison))

        container = QWidget()
        container.setLayout(layout)
        self.setCentralWidget(container)

    def create_button(self, text, callback):
        """Create a QPushButton with a callback."""
        button = QPushButton(text)
        button.clicked.connect(callback)
        return button

    def select_file(self, label, file_attr):
        """Generic method to select a file."""
        file_path, _ = QFileDialog.getOpenFileName(self, "Select BOM File", ".", "Excel Files (*.xlsx)")
        if file_path:
            setattr(self, file_attr, file_path)
            label.setText(f"{file_attr.replace('file', 'BOM File')}: {os.path.basename(file_path)}")

    def select_file1(self):
        self.select_file(self.label_file1, 'file1')

    def select_file2(self):
        self.select_file(self.label_file2, 'file2')

    def run_comparison(self):
        """Run the BOM comparison."""
        if not self.file1 or not self.file2:
            QMessageBox.warning(self, "Error", "Please select both BOM files before running the comparison.")
            return

        try:
            os.makedirs(OUTPUT_DIR, exist_ok=True)
            output_file = self.get_output_file_name()

            df1, df2 = self.read_excel_files()
            df_merged = self.merge_and_filter_dataframes(df1, df2)

            output_file = self.handle_existing_file(output_file)
            if not output_file:
                return

            self.write_to_excel(output_file, df1, df2, df_merged)
            self.modify_excel_file(output_file)
            self.write_refdes_to_txt(df_merged)

            QMessageBox.information(self, "Success", f"Differences saved to {output_file}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"An error occurred: {e}")

    def get_output_file_name(self):
        """Generate the output file name."""
        bom1_name = os.path.basename(self.file1).split('.')[0].split('_')[-1]
        bom2_name = os.path.basename(self.file2).split('.')[0].split('_')[-1]
        return os.path.join(OUTPUT_DIR, f'BOM_Differences_{bom1_name}_vs_{bom2_name}.xlsx')

    def read_excel_files(self):
        """Read the Excel files into DataFrames."""
        df1 = pd.read_excel(self.file1, skiprows=15, skipfooter=7, usecols=COLUMNS).set_index('RefDes')
        df2 = pd.read_excel(self.file2, skiprows=15, skipfooter=7, usecols=COLUMNS).set_index('RefDes')
        return df1, df2

    def merge_and_filter_dataframes(self, df1, df2):
        """Merge and apply filters to the DataFrames."""
        df_merged = pd.merge(df1, df2, how='outer', left_index=True, right_index=True, suffixes=('_BOM1', '_BOM2'))

        if self.checkbox_remove_zero_qty.isChecked():
            df_merged = df_merged[~((df_merged['Quantity_BOM1'] == 0) & (df_merged['Quantity_BOM2'] == 0))]

        if self.checkbox_remove_new_zero_qty.isChecked():
            df_merged = df_merged[~((df_merged['Quantity_BOM2'] == 0) & (df_merged['Quantity_BOM1'].isna()))]

        if self.checkbox_remove_deprecated_zero_qty.isChecked():
            df_merged = df_merged[~((df_merged['Quantity_BOM1'] == 0) & (df_merged['Quantity_BOM2'].isna()))]

        if self.checkbox_remove_unchanged.isChecked():
            df_merged = df_merged.loc[~(df_merged.filter(like='_BOM1').fillna('').values ==
                                        df_merged.filter(like='_BOM2').fillna('').values).all(axis=1)]

        return df_merged.sort_index()

    def handle_existing_file(self, output_file):
        """Handle cases where the output file already exists."""
        if os.path.exists(output_file):
            msg_box = QMessageBox(self)
            msg_box.setWindowTitle("File Already Exists")
            msg_box.setText(f"The file '{os.path.basename(output_file)}' already exists.")
            msg_box.setInformativeText("What would you like to do?")
            overwrite_button = msg_box.addButton("Overwrite", QMessageBox.AcceptRole)
            rename_button = msg_box.addButton("Rename", QMessageBox.ActionRole)
            cancel_button = msg_box.addButton("Cancel", QMessageBox.RejectRole)
            msg_box.exec_()

            clicked_button = msg_box.clickedButton()
            if clicked_button == overwrite_button:
                return output_file
            elif clicked_button == rename_button:
                timestamp = time.strftime("%d.%m.%y_%H%M")
                return os.path.join(OUTPUT_DIR, f'{os.path.basename(output_file).split(".")[0]}_{timestamp}.xlsx')
            elif clicked_button == cancel_button:
                QMessageBox.information(self, "Operation Cancelled", "The operation has been cancelled.")
                return None
        return output_file

    def write_to_excel(self, output_file, df1, df2, df_merged):
        """Write the DataFrames to an Excel file."""
        while True:
            try:
                with pd.ExcelWriter(output_file, mode='w') as writer:
                    pd.concat([df1, df2], axis=1, join='outer').sort_index().to_excel(writer, sheet_name='Side by Side Comparison')
                    df_merged.to_excel(writer, sheet_name='BOM Delta')
                break
            except PermissionError:
                if not self.handle_permission_error():
                    return

    def handle_permission_error(self):
        """Handle permission errors when writing to a file."""
        retry_msg = QMessageBox(self)
        retry_msg.setWindowTitle("Permission Error")
        retry_msg.setText("The file is open or being used by another process.")
        retry_msg.setInformativeText("Please close the file and try again.")
        retry_button = retry_msg.addButton("Retry", QMessageBox.AcceptRole)
        cancel_button = retry_msg.addButton("Cancel", QMessageBox.RejectRole)
        retry_msg.exec_()

        if retry_msg.clickedButton() == cancel_button:
            QMessageBox.information(self, "Operation Cancelled", "The operation has been cancelled.")
            return False
        time.sleep(1)
        return True

    def modify_excel_file(self, output_file):
        """Modify the Excel file using openpyxl."""
        wb = load_workbook(output_file)
        ws = wb['Side by Side Comparison']

        ws.insert_rows(1)
        ws.merge_cells('B1:I1')
        ws.merge_cells('J1:Q1')
        ws['B1'], ws['J1'] = 'BOM1', 'BOM2'

        bold_font = Font(bold=True)
        center_alignment = Alignment(horizontal='center')
        for cell in ['B1', 'J1']:
            ws[cell].font = bold_font
            ws[cell].alignment = center_alignment

        ws.auto_filter.ref = 'A2:U2'
        wb.save(output_file)
        os.startfile(output_file)

    def write_refdes_to_txt(self, df_merged):
        """Write the RefDes list to a text file."""
        bom1_name = os.path.basename(self.file1).split('.')[0]
        bom2_name = os.path.basename(self.file2).split('.')[0]
        txt_file = os.path.join(OUTPUT_DIR, f'RefDesList_{bom1_name}_vs_{bom2_name}.txt')

        with open(txt_file, 'w') as f:
            f.writelines(f"{refdes}\n" for refdes in df_merged.index)


def main():
    app = QApplication(sys.argv)
    window = BOMComparisonApp()
    window.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
